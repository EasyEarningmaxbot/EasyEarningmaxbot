import io
import openpyxl
import logging
import asyncio
from aiogram import Router, types, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder, InlineKeyboardButton, ReplyKeyboardBuilder, KeyboardButton

import config
from database import (
    users_col, tasks_col, withdrawals_col, get_setting, update_setting, 
    update_balance, set_banned, is_task_type_active, toggle_task_type_status
)
from keyboards import (
    admin_menu, main_menu, cancel_keyboard, 
    reports_menu, user_money_menu, task_settings_menu, task_on_off_menu, others_menu
)

router = Router()

# -------------------------------------------------------------
# 🛠️ FSM States (Next Step Handler এর সম্পূর্ণ বিকল্প)
# -------------------------------------------------------------
class AdminStates(StatesGroup):
    waiting_fb_file = State()
    waiting_ig_file = State()
    waiting_pass = State()
    waiting_price = State()
    waiting_leaderboard_prizes = State()
    waiting_announcement = State()
    waiting_ban_id = State()
    waiting_bal_id = State()
    waiting_bal_amount = State()
    waiting_method_limit = State()
    waiting_ref_comm = State()
    waiting_video_link = State()
    confirm_delete_tasks = State()

def is_admin(user_id: int) -> bool:
    return user_id in config.ADMIN_IDS

# -------------------------------------------------------------
# 📌 ১. অ্যাডমিন প্যানেল ও মেনু নেভিগেশন
# -------------------------------------------------------------
@router.message(F.text == 'অ্যাডমিন প্যানেল')
async def handle_admin_panel(message: types.Message):
    if is_admin(message.from_user.id):
        markup = await admin_menu() if callable(admin_menu) else admin_menu
        await message.answer("অ্যাডমিন প্যানেলে স্বাগতম!", reply_markup=markup)
    else:
        await message.answer("অনুমতি নেই।")

@router.message(F.text == 'রিপোর্ট')
async def handle_reports_category(message: types.Message):
    if not is_admin(message.from_user.id): return
    markup = await reports_menu() if callable(reports_menu) else reports_menu
    await message.answer("<b>Reports Panel</b>", reply_markup=markup)

@router.message(F.text == 'ইউজার ও টাকা')
async def handle_user_money_category(message: types.Message):
    if not is_admin(message.from_user.id): return
    markup = await user_money_menu() if callable(user_money_menu) else user_money_menu
    await message.answer("<b>ইউজার ও টাকা সেটিংস</b>", reply_markup=markup)

@router.message(F.text == 'টাস্ক সেটিংস')
async def handle_task_settings_category(message: types.Message):
    if not is_admin(message.from_user.id): return
    markup = await task_settings_menu() if callable(task_settings_menu) else task_settings_menu
    await message.answer("<b>টাস্ক সেটিংস</b>", reply_markup=markup)

@router.message(F.text == 'অন্যান্য')
async def handle_others_category(message: types.Message):
    if not is_admin(message.from_user.id): return
    markup = await others_menu() if callable(others_menu) else others_menu
    await message.answer("<b>অন্যান্য সেটিংস</b>", reply_markup=markup)

@router.message(F.text == 'প্রধান মেনু')
async def back_to_main_menu(message: types.Message):
    markup = await main_menu(message.from_user.id) if callable(main_menu) else main_menu
    await message.answer("প্রধান মেনুতে ফিরে যাচ্ছি...", reply_markup=markup)

# -------------------------------------------------------------
# 📊 ২. টাস্ক রিপোর্ট (Excel Export)
# -------------------------------------------------------------
@router.message(F.text == 'টাস্ক রিপোর্ট (Excel)')
async def download_tasks_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text='📥 IG 2FA [Excel Report]'))
    builder.row(KeyboardButton(text='📥 FB Cookies [Excel Report]'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer("<b>কোন টাস্কের এক্সেল ফাইল নামাতে চান সিলেক্ট করুন:</b>", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text == '📥 IG 2FA [Excel Report]')
async def export_instagram_excel(message: types.Message):
    if not is_admin(message.from_user.id): return
    
    def generate_excel():
        pending_tasks = list(tasks_col.find({
            '$or': [{'type': 'Instagram'}, {'type': {'$exists': False}}], 
            'status': 'Pending Review',
            'username': {'$exists': True, '$ne': ''}
        }))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instagram Tasks"
        ws.append(['Username', 'Password', '2FA Key'])
        
        valid_tasks = [t for t in pending_tasks if t.get('username')]
        if not valid_tasks:
            ws.append(['No data', 'No data', 'No data'])
        else:
            for task in valid_tasks:
                ws.append([task.get('username', ''), task.get('password', ''), task.get('2fa_key', '')])
                
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    excel_bytes = await asyncio.to_thread(generate_excel)
    file_bytes = types.BufferedInputFile(excel_bytes, filename="Instagram_Task_Report.xlsx")
    markup = await reports_menu() if callable(reports_menu) else reports_menu
    await message.answer_document(file_bytes, caption="<b>ইন্সটাগ্রাম কাজের পেন্ডিং Excel ফাইল।</b>", reply_markup=markup)

@router.message(F.text == '📥 FB Cookies [Excel Report]')
async def export_facebook_excel(message: types.Message):
    if not is_admin(message.from_user.id): return
    
    def generate_excel():
        pending_tasks = list(tasks_col.find({
            'type': 'Facebook', 
            'status': 'Pending Review',
            'fb_uid': {'$exists': True, '$ne': ''}
        }))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facebook Tasks"
        ws.append(['UID', 'Password', 'Cookies'])
        
        default_fb_pass = get_setting('fb_password') or get_setting('password') or 'kamrol@22'
        valid_tasks = [t for t in pending_tasks if t.get('fb_uid')]
        if not valid_tasks:
            ws.append(['No data', 'No data', 'No data'])
        else:
            for task in valid_tasks:
                ws.append([task.get('fb_uid', ''), task.get('password', default_fb_pass), task.get('fb_cookie', '')])
                
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    excel_bytes = await asyncio.to_thread(generate_excel)
    file_bytes = types.BufferedInputFile(excel_bytes, filename="Facebook_Task_Report.xlsx")
    markup = await reports_menu() if callable(reports_menu) else reports_menu
    await message.answer_document(file_bytes, caption="<b>ফেসবুক কাজের পেন্ডিং Excel ফাইল।</b>", reply_markup=markup)

# -------------------------------------------------------------
# 📥 ৩. রিপোর্ট সাবমিট (Excel Import Accept/Reject)
# -------------------------------------------------------------
@router.message(F.text == 'রিপোর্ট সাবমিট')
async def report_submit_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text='📤 FB Cookies [Submit Report]'))
    builder.row(KeyboardButton(text='📤 IG 2FA [Submit Report]'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer("<b>রিপোর্ট সাবমিট এর জন্য টাস্ক নির্বাচন করুন:</b>", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text.in_(['📤 FB Cookies [Submit Report]', '📤 IG 2FA [Submit Report]']))
async def submit_action_choice(message: types.Message):
    if not is_admin(message.from_user.id): return
    task_type = 'Facebook' if 'FB' in message.text else 'Instagram'
    builder = ReplyKeyboardBuilder()
    if task_type == 'Facebook':
        builder.row(KeyboardButton(text='FB Accept Tasks'), KeyboardButton(text='FB Reject Tasks'))
    else:
        builder.row(KeyboardButton(text='IG Accept Tasks'), KeyboardButton(text='IG Reject Tasks'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer(f"<b>{message.text}</b> এর জন্য অপশন নির্বাচন করুন:", reply_markup=builder.as_markup(resize_keyboard=True))

# FB Import Process
@router.message(F.text.in_(['FB Accept Tasks', 'FB Reject Tasks']))
async def trigger_fb_file_input(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    action = 'accept' if 'Accept' in message.text else 'reject'
    await state.update_data(action=action)
    
    msg_text = (
        '⚙️ 📘 <b>Facebook (Number) — ' + ('Accept' if action == 'accept' else 'Reject') + ' করতে .xlsx ফাইল পাঠান।</b>\n\n'
        'Format: (Column A: UID)\n(exported file থেকে first column select করে upload করো, header row থাকবে)'
    )
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(msg_text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_fb_file)

@router.message(AdminStates.waiting_fb_file, F.document)
async def process_facebook_excel_file(message: types.Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    action = data.get('action')
    await state.clear()
    
    file_bytes = await bot.download(message.document)
    
    def process_file():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes.read()))
        sheet = wb.active
        uids = [str(row[0]).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row and row[0]]
        
        fb_price = get_setting('fb_price') or 0.00
        ref_percent = get_setting('ref_commission') or 10
        user_summary = {}
        processed_count = 0

        for uid in uids:
            task = tasks_col.find_one({'fb_uid': uid, 'type': 'Facebook', 'status': 'Pending Review'})
            if task:
                w_id = task['user_id']
                user_summary[w_id] = user_summary.get(w_id, 0) + 1
                tasks_col.delete_one({'_id': task['_id']})
                processed_count += 1

        return user_summary, processed_count, fb_price, ref_percent

    user_summary, processed_count, fb_price, ref_percent = await asyncio.to_thread(process_file)

    for worker_id, count in user_summary.items():
        total_earned = count * fb_price
        if action == 'accept':
            await asyncio.to_thread(update_balance, worker_id, total_earned)
            await asyncio.to_thread(users_col.update_one, {'user_id': worker_id}, {'$inc': {'total_income': total_earned, 'completed_tasks': count}})
            
            try:
                await bot.send_message(worker_id, f"🎉 <b>অভিনন্দন আপনার কাজটা এপ্রুভ করা হয়েছে এবং আপনি কাজের জন্য ({total_earned:.2f})টাকা পেয়েছেন</b> 📘")
            except Exception: pass

            user_data = await asyncio.to_thread(users_col.find_one, {'user_id': worker_id})
            referrer_id = user_data.get('referrer_id') if user_data else None
            if referrer_id:
                comm_amount = total_earned * (ref_percent / 100.0)
                if comm_amount > 0:
                    await asyncio.to_thread(update_balance, referrer_id, comm_amount)
                    await asyncio.to_thread(users_col.update_one, {'user_id': referrer_id}, {'$inc': {'total_income': comm_amount, 'ref_income': comm_amount}})
                    try:
                        await bot.send_message(referrer_id, f"🎁 <b>আপনার রেফার থেকে আপনি {comm_amount:.2f} টাকা পেয়েছেন অভিনন্দন আপনি আরো রেফার করুন</b> 🚀")
                    except Exception: pass
        else:
            try:
                await bot.send_message(worker_id, f"❌ <b>সরি ভাইয়া আপনার এই {count} টা একাউন্ট ব্যান হয়ে গিয়েছে বায়ার নেয় নাই এটার জন্য আপনি কোন পেমেন্ট পাবেন না</b> 📘")
            except Exception: pass

    markup = await reports_menu() if callable(reports_menu) else reports_menu
    await message.answer(f"<b>ফেসবুক ফাইল প্রসেস সম্পন্ন!</b>\nমোট আইটেম প্রসেসড: {processed_count}", reply_markup=markup)

# IG Import Process
@router.message(F.text.in_(['IG Accept Tasks', 'IG Reject Tasks']))
async def trigger_ig_file_input(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    action = 'accept' if 'Accept' in message.text else 'reject'
    await state.update_data(action=action)
    
    msg_text = (
        '✨ 📱 <b>Instagram 2FA — ' + ('Accept' if action == 'accept' else 'Reject') + ' করতে .xlsx ফাইল পাঠান।</b>\n\n'
        'Format: (Column A: Username)\n(exported file থেকে first column select করে upload করো, header row থাকবে)'
    )
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(msg_text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_ig_file)

@router.message(AdminStates.waiting_ig_file, F.document)
async def process_instagram_excel_file(message: types.Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    action = data.get('action')
    await state.clear()
    
    file_bytes = await bot.download(message.document)
    
    def process_file():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes.read()))
        sheet = wb.active
        usernames = [str(row[0]).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row and row[0]]
        
        ig_price = get_setting('price') or 2.70
        ref_percent = get_setting('ref_commission') or 10
        user_summary = {}
        processed_count = 0

        for uname in usernames:
            task = tasks_col.find_one({'username': uname, 'status': 'Pending Review'})
            if task:
                w_id = task['user_id']
                user_summary[w_id] = user_summary.get(w_id, 0) + 1
                tasks_col.delete_one({'_id': task['_id']})
                processed_count += 1

        return user_summary, processed_count, ig_price, ref_percent

    user_summary, processed_count, ig_price, ref_percent = await asyncio.to_thread(process_file)

    for worker_id, count in user_summary.items():
        total_earned = count * ig_price
        if action == 'accept':
            await asyncio.to_thread(update_balance, worker_id, total_earned)
            await asyncio.to_thread(users_col.update_one, {'user_id': worker_id}, {'$inc': {'total_income': total_earned, 'completed_tasks': count}})
            
            try:
                await bot.send_message(worker_id, f"🎉 <b>বস কি খেলা দেখাইলা তোমার কাজ এপ্রোভ হয়েছে তোমার ব্যালেন্সে {total_earned:.2f} টাকা এড করা হয়েছে,</b> 📱")
            except Exception: pass

            user_data = await asyncio.to_thread(users_col.find_one, {'user_id': worker_id})
            referrer_id = user_data.get('referrer_id') if user_data else None
            if referrer_id:
                comm_amount = total_earned * (ref_percent / 100.0)
                if comm_amount > 0:
                    await asyncio.to_thread(update_balance, referrer_id, comm_amount)
                    await asyncio.to_thread(users_col.update_one, {'user_id': referrer_id}, {'$inc': {'total_income': comm_amount, 'ref_income': comm_amount}})
                    try:
                        await bot.send_message(referrer_id, f"🎁 <b>আপনার রেফার থেকে আপনি {comm_amount:.2f} টাকা পেয়েছেন অভিনন্দন আপনি আরো রেফার করুন</b> 🚀")
                    except Exception: pass
        else:
            try:
                await bot.send_message(worker_id, f"❌ <b>দুঃখিত আপনি এই অ্যাকাউন্ট খোলার জন্য পেমেন্ট পাবেন না কারণ আপনাদের একাউন্টগুলা বেন হয়ে গেছে</b> 📱")
            except Exception: pass

    markup = await reports_menu() if callable(reports_menu) else reports_menu
    await message.answer(f"<b>ইন্সটাগ্রাম ফাইল প্রসেস সম্পন্ন!</b>\nমোট আইটেম প্রসেসড: {processed_count}", reply_markup=markup)

# -------------------------------------------------------------
# 🗑️ ৪. সব পেন্ডিং কাজ রিমুভ
# -------------------------------------------------------------
@router.message(F.text == 'সব পেন্ডিং কাজ রিমুভ')
async def prompt_delete_all_pending_tasks(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    pending_count = await asyncio.to_thread(tasks_col.count_documents, {'status': 'Pending Review'})
    
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text='হ্যাঁ, ডিলিট করুন'), KeyboardButton(text='বাতিল'))
    
    text = (
        f'<b>সতর্কতা!</b>\n\n'
        f'বর্তমানে মোট <b>{pending_count}টি</b> পেন্ডিং কাজ রিভিউ তালিকায় রয়েছে।\n'
        f'আপনি কি নিশ্চিত যে রিভিউতে থাকা <b>সব কাজ ডাটাবেজ থেকে মুছে ফেলতে চান?</b>'
    )
    await message.answer(text, reply_markup=builder.as_markup(resize_keyboard=True))
    await state.set_state(AdminStates.confirm_delete_tasks)

@router.message(AdminStates.confirm_delete_tasks, F.text == 'হ্যাঁ, ডিলিট করুন')
async def process_delete_all_pending_tasks(message: types.Message, state: FSMContext):
    await state.clear()
    result = await asyncio.to_thread(tasks_col.delete_many, {'status': 'Pending Review'})
    markup = await task_settings_menu() if callable(task_settings_menu) else task_settings_menu
    await message.answer(f"<b>সফলভাবে রিভিউতে থাকা সকল ({result.deleted_count}টি) কাজ ডাটাবেজ থেকে মুছে ফেলা হয়েছে!</b>", reply_markup=markup)

# -------------------------------------------------------------
# 🔑 ৫. পাসওয়ার্ড সেট (ইন্সটাগ্রাম ও ফেসবুক)
# -------------------------------------------------------------
@router.message(F.text == 'পাসওয়ার্ড সেট')
async def set_task_password_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text='Instagram 2FA Password'))
    builder.row(KeyboardButton(text='Anymail/Number Password'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer("🔑 <b>কোন টাস্কের পাসওয়ার্ড সেট করতে চান:</b>", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text.in_(['Instagram 2FA Password', 'Anymail/Number Password']))
async def prompt_pass_change(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    is_fb = ('Anymail' in message.text)
    setting_key = 'fb_password' if is_fb else 'password'
    curr_pass = await asyncio.to_thread(get_setting, setting_key) or 'kamrol@22'
    
    await state.update_data(setting_key=setting_key)
    msg_text = f"<b>বর্তমান পাসওয়ার্ড:</b> <code>{curr_pass}</code>\n<b>নতুন পাসওয়ার্ড দিন:</b>"
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(msg_text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_pass)

@router.message(AdminStates.waiting_pass, F.text != 'বাতিল')
async def process_save_pass(message: types.Message, state: FSMContext):
    data = await state.get_data()
    setting_key = data.get('setting_key')
    await state.clear()
    
    new_pass = message.text.strip()
    await asyncio.to_thread(update_setting, setting_key, new_pass)
    markup = await others_menu() if callable(others_menu) else others_menu
    await message.answer(f"পাসওয়ার্ড সফলভাবে আপডেট করা হয়েছে: <code>{new_pass}</code>", reply_markup=markup)

# -------------------------------------------------------------
# 💰 ৬. কাজের মূল্য সেট
# -------------------------------------------------------------
@router.message(F.text == 'কাজের মূল্য সেট')
async def set_task_price_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text='IG কাজের মূল্য সেট'))
    builder.row(KeyboardButton(text='FB কাজের মূল্য সেট'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer("<b>কোন কাজের মূল্য সেট করতে চান তা সিলেক্ট করুন:</b>", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text.in_(['IG কাজের মূল্য সেট', 'FB কাজের মূল্য সেট']))
async def prompt_price_change(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    is_fb = ('FB' in message.text)
    setting_key = 'fb_price' if is_fb else 'price'
    curr_price = await asyncio.to_thread(get_setting, setting_key) or 0.00
    
    await state.update_data(setting_key=setting_key)
    msg_text = f"<b>বর্তমান মূল্য আছে {curr_price:.2f}৳। আপনার নতুন মূল্যটি নিচে লিখে দিন:</b>"
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(msg_text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_price)

@router.message(AdminStates.waiting_price, F.text != 'বাতিল')
async def process_save_price(message: types.Message, state: FSMContext):
    data = await state.get_data()
    setting_key = data.get('setting_key')
    await state.clear()
    
    markup = await user_money_menu() if callable(user_money_menu) else user_money_menu
    try:
        new_price = float(message.text.strip())
        await asyncio.to_thread(update_setting, setting_key, new_price)
        await message.answer("<b>অভিনন্দন! আপনার মূল্যটি চেঞ্জ করা হয়েছে 🎉</b>", reply_markup=markup)
    except ValueError:
        await message.answer("<b>সঠিক সংখ্যা দিয়ে মূল্য পুনরায় চেষ্টা করুন।</b>", reply_markup=markup)

# -------------------------------------------------------------
# 🏆 ৭. লিডারবোর্ড কন্ট্রোল
# -------------------------------------------------------------
@router.message(F.text == 'লিডার বোর্ড ON OF')
async def toggle_leaderboard(message: types.Message):
    if not is_admin(message.from_user.id): return
    current_status = await asyncio.to_thread(get_setting, 'leaderboard_active')
    if current_status is None: current_status = True
        
    new_status = not current_status
    await asyncio.to_thread(update_setting, 'leaderboard_active', new_status)
    
    text = "📍 <b>লিডার বোর্ড এই মুহূর্তে ON করা হলো</b>" if new_status else "📍 <b>লিডার বোর্ড এই মুহূর্তে OFF করা হলো</b>"
    markup = await others_menu() if callable(others_menu) else others_menu
    await message.answer(text, reply_markup=markup)

@router.message(F.text == 'লিডার বোর্ড প্রাইস সেট')
async def set_leaderboard_prizes_prompt(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    curr_prizes = await asyncio.to_thread(get_setting, 'leaderboard_prizes') or "100,50,30,20,10"
    text = f"🎁 <b>টপ ৫ র‍্যাংকের জন্য reward amount (৳) লিখুন (কমা দিয়ে আলাদা করে):</b>\n\n<b>বর্তমান:</b> {curr_prizes}"
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_leaderboard_prizes)

@router.message(AdminStates.waiting_leaderboard_prizes, F.text != 'বাতিল')
async def process_save_leaderboard_prizes(message: types.Message, state: FSMContext):
    await state.clear()
    raw_text = message.text.strip().strip(',')
    p_list = [p.strip() for p in raw_text.split(',') if p.strip()]
    markup = await others_menu() if callable(others_menu) else others_menu
    
    if len(p_list) == 5:
        try:
            _ = [float(p) for p in p_list]
            clean_prize_str = ",".join(p_list)
            await asyncio.to_thread(update_setting, 'leaderboard_prizes', clean_prize_str)
            await message.answer(f"<b>লিডারবোর্ড প্রাইজ সফলভাবে আপডেট হয়েছে!</b>\n<b>বর্তমান প্রাইজ:</b> {clean_prize_str}", reply_markup=markup)
        except ValueError:
            await message.answer("<b>ভুল ইনপুট! শুধু সংখ্যা ও কমা ব্যবহার করুন।</b>", reply_markup=markup)
    else:
        await message.answer("<b>অবশ্যই ৫ জন র‍্যাংকের জন্য কমা দিয়ে আলাদা করে অ্যামাউন্ট লিখুন।</b>", reply_markup=markup)

# -------------------------------------------------------------
# 📢 ৮. এনাউন্সমেন্ট / ব্রডকাস্ট
# -------------------------------------------------------------
@router.message(F.text == 'এনাউন্সমেন্ট')
async def admin_announcement(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer("বটের সকল মেম্বারদের কাছে যে অফিশিয়াল নোটিশটি পাঠাতে চান, তা টাইপ করুন:", reply_markup=markup)
    await state.set_state(AdminStates.waiting_announcement)

@router.message(AdminStates.waiting_announcement, F.text != 'বাতিল')
async def process_announcement(message: types.Message, state: FSMContext, bot: Bot):
    await state.clear()
    announcement_text = message.text
    all_users = await asyncio.to_thread(lambda: list(users_col.find({}, {'user_id': 1})))
    
    markup = await others_menu() if callable(others_menu) else others_menu
    await message.answer(f"এনাউন্সমেন্ট পাঠানো শুরু হয়েছে... (মোট ইউজার: {len(all_users)} জন)", reply_markup=markup)
    
    async def send_msg(user_id):
        try:
            await bot.send_message(user_id, f"<b>অফিশিয়াল ঘোষণা:</b>\n\n{announcement_text}")
            return True
        except Exception: return False

    tasks = [send_msg(u['user_id']) for u in all_users]
    results = await asyncio.gather(*tasks)
    success_count = sum(1 for r in results if r)
    await message.answer(f"নোটিশ পাঠানো সফল হয়েছে!\nমোট {success_count} জন সচল মেম্বার ইনবক্সে মেসেজটি পেয়েছে।", reply_markup=markup)

# -------------------------------------------------------------
# 🔎 ৯. উত্তোলন রিভিউ, ব্যান/আনব্যান ও ব্যালেন্স কন্ট্রোল
# -------------------------------------------------------------
@router.message(F.text == 'উত্তোলন রিভিউ')
async def view_withdraw_requests(message: types.Message):
    if not is_admin(message.from_user.id): return
    pending = await asyncio.to_thread(lambda: list(withdrawals_col.find({'status': 'Pending'})))
    if not pending:
        await message.answer("কোনো পেন্ডিং উইথড্র রিকোয়েস্ট নেই।")
        return
    for w in pending[:5]:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="Approve", callback_data=f"gr_wd_approve_{w['_id']}"),
            InlineKeyboardButton(text="Reject", callback_data=f"gr_wd_reject_{w['_id']}")
        )
        text = f"<b>উইথড্র রিকোয়েস্ট রিভিউ (বক্স):</b>\nইউজার: {w['first_name']}\nমেথড: {w['method']}\nনম্বর: {w['number']}\nপরিমাণ: ৳{w['amount']:.2f}"
        await message.answer(text, reply_markup=builder.as_markup())

@router.message(F.text == 'ব্যান / আনব্যান')
async def ban_unban_user(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer("ব্যান বা আনব্যান করতে ইউজারের ID দিন:", reply_markup=markup)
    await state.set_state(AdminStates.waiting_ban_id)

@router.message(AdminStates.waiting_ban_id, F.text != 'বাতিল')
async def process_ban_unban(message: types.Message, state: FSMContext):
    await state.clear()
    markup = await task_settings_menu() if callable(task_settings_menu) else task_settings_menu
    try:
        target_id = int(message.text.strip())
        user = await asyncio.to_thread(users_col.find_one, {'user_id': target_id})
        if user:
            new_status = not user.get('is_banned', False)
            await asyncio.to_thread(set_banned, target_id, new_status)
            await message.answer(f"ইউজার {target_id} এর স্ট্যাটাস আপডেট সফল!", reply_markup=markup)
    except Exception: pass

@router.message(F.text == 'ব্যালেন্স অ্যাড/রিমুভ')
async def balance_add_remove(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer("ইউজারের Telegram ID দিন:", reply_markup=markup)
    await state.set_state(AdminStates.waiting_bal_id)

@router.message(AdminStates.waiting_bal_id, F.text != 'বাতিল')
async def process_bal_change_id(message: types.Message, state: FSMContext):
    try:
        t_id = int(message.text.strip())
        await state.update_data(target_id=t_id)
        markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
        await message.answer("টাকার পরিমাণ লিখুন (কেটে নিতে মাইনাস সহ):", reply_markup=markup)
        await state.set_state(AdminStates.waiting_bal_amount)
    except Exception:
        await state.clear()

@router.message(AdminStates.waiting_bal_amount, F.text != 'বাতিল')
async def process_bal_change_amount(message: types.Message, state: FSMContext):
    data = await state.get_data()
    t_id = data.get('target_id')
    await state.clear()
    markup = await user_money_menu() if callable(user_money_menu) else user_money_menu
    try:
        val = float(message.text)
        await asyncio.to_thread(update_balance, t_id, val)
        if val > 0:
            await asyncio.to_thread(users_col.update_one, {'user_id': t_id}, {'$inc': {'total_income': val}})
        await message.answer("ব্যালেন্স সফলভাবে আপডেট করা হয়েছে।", reply_markup=markup)
    except Exception: pass

# -------------------------------------------------------------
# ⚙️ ১০. কাজ ON/OFF
# -------------------------------------------------------------
@router.message(F.text == 'কাজ On/Off')
async def open_task_on_off_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    markup = await task_on_off_menu() if callable(task_on_off_menu) else task_on_off_menu
    await message.answer("<b>যেই কাজটি অন বা অফ করতে চান সেটির ওপর ক্লিক করুন:</b>", reply_markup=markup)

@router.message(F.text.contains('ইন্সটাগ্রাম 2FA') | F.text.contains('0 fnd cookies'))
async def toggle_specific_task_status(message: types.Message):
    if not is_admin(message.from_user.id): return
    task_type = 'Instagram' if 'ইন্সটাগ্রাম' in message.text else 'Facebook'
    new_status = await asyncio.to_thread(toggle_task_type_status, task_type)
    status_text = "ON" if new_status else "OFF"
    markup = await task_on_off_menu() if callable(task_on_off_menu) else task_on_off_menu
    await message.answer(f"<b>{task_type} কাজ এখন [{status_text}] করা হয়েছে।</b>", reply_markup=markup)

# -------------------------------------------------------------
# 💳 ১১. উত্তোলন মেথড ON/OFF
# -------------------------------------------------------------
@router.message(F.text.in_(['উত্তোলনের মেথড On/Off', 'উত্তোলন মেথড On/Off']))
async def handle_withdraw_toggle_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    await show_toggle_keyboard(message)

async def show_toggle_keyboard(message: types.Message):
    bkash_status = await asyncio.to_thread(get_setting, 'status_bkash')
    nagad_status = await asyncio.to_thread(get_setting, 'status_nagad')
    recharge_status = await asyncio.to_thread(get_setting, 'status_recharge')
    usdt_status = await asyncio.to_thread(get_setting, 'status_usdt')

    bkash_status = bkash_status if bkash_status is not None else True
    nagad_status = nagad_status if nagad_status is not None else True
    recharge_status = recharge_status if recharge_status is not None else True
    usdt_status = usdt_status if usdt_status is not None else True

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text=f"বিকাশ: {'ON' if bkash_status else 'OFF'}", callback_data="toggle_wd_bkash"))
    builder.row(InlineKeyboardButton(text=f"নগদ: {'ON' if nagad_status else 'OFF'}", callback_data="toggle_wd_nagad"))
    builder.row(InlineKeyboardButton(text=f"মোবাইল রিচার্জ: {'ON' if recharge_status else 'OFF'}", callback_data="toggle_wd_recharge"))
    builder.row(InlineKeyboardButton(text=f"USDT (BEP-20): {'ON' if usdt_status else 'OFF'}", callback_data="toggle_wd_usdt"))

    await message.answer("<b>উত্তোলন মেথড অন/অফ কন্ট্রোল প্যানেল:</b>\nনিচের বাটনে ক্লিক করে অন বা অফ করুন।", reply_markup=builder.as_markup())

@router.callback_query(F.data.startswith('toggle_wd_'))
async def handle_withdraw_toggle_callback(call: types.CallbackQuery):
    if not is_admin(call.from_user.id): return
    method = call.data.split('_')[2]
    key = f'status_{method}'
    
    current_status = await asyncio.to_thread(get_setting, key)
    current_status = current_status if current_status is not None else True
    new_status = not current_status
    
    await asyncio.to_thread(update_setting, key, new_status)
    status_text = "চালু (ON)" if new_status else "বন্ধ (OFF)"
    await call.answer(f"{method.upper()} মেথডটি {status_text} করা হয়েছে!", show_alert=True)
    
    try: await call.message.delete()
    except Exception: pass
    await show_toggle_keyboard(call.message)

# -------------------------------------------------------------
# 💸 ১২. উত্তোলন লিমিট সেট
# -------------------------------------------------------------
@router.message(F.text.in_(['উত্তোলনের লিমিট', 'উত্তোলন লিমিট']))
async def withdraw_limit_menu(message: types.Message):
    if not is_admin(message.from_user.id): return
    usdt_limit = await asyncio.to_thread(get_setting, 'min_withdraw_usdt') or 25.00
    bkash_limit = await asyncio.to_thread(get_setting, 'min_withdraw_bkash') or 100.00
    nagad_limit = await asyncio.to_thread(get_setting, 'min_withdraw_nagad') or 100.00
    recharge_limit = await asyncio.to_thread(get_setting, 'min_withdraw_recharge') or 20.00

    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text=f'USDT (BEP-20) -> লিমিট {usdt_limit:.2f}(~0.05)'))
    builder.row(KeyboardButton(text=f'বিকাশ -> লিমিট {bkash_limit:.2f}৳(~৫)'))
    builder.row(KeyboardButton(text=f'নগদ -> লিমিট {nagad_limit:.2f}৳(~৫)'))
    builder.row(KeyboardButton(text=f'মোবাইল রিচার্জ -> লিমিট {recharge_limit:.2f}৳'))
    builder.row(KeyboardButton(text='অ্যাডমিন প্যানেল'))
    await message.answer("<b>কোন মেথডের উত্তোলন লিমিট পরিবর্তন করতে চান সিলেক্ট করুন:</b>", reply_markup=builder.as_markup(resize_keyboard=True))

@router.message(F.text.contains('USDT (BEP-20)') | F.text.contains('বিকাশ') | F.text.contains('নগদ') | F.text.contains('মোবাইল রিচার্জ'))
async def prompt_method_limit_change(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    text = message.text
    if 'USDT' in text:
        method_key = 'min_withdraw_usdt'
        curr_limit = await asyncio.to_thread(get_setting, method_key) or 25.00
    elif 'বিকাশ' in text:
        method_key = 'min_withdraw_bkash'
        curr_limit = await asyncio.to_thread(get_setting, method_key) or 100.00
    elif 'নগদ' in text:
        method_key = 'min_withdraw_nagad'
        curr_limit = await asyncio.to_thread(get_setting, method_key) or 100.00
    else:
        method_key = 'min_withdraw_recharge'
        curr_limit = await asyncio.to_thread(get_setting, method_key) or 20.00

    await state.update_data(method_key=method_key)
    prompt_text = f"বর্তমানে সর্বনিম্ন সেট করা আছে ({curr_limit:.2f})। নতুন লিমিট নিচে লিখুন:"
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer(prompt_text, reply_markup=markup)
    await state.set_state(AdminStates.waiting_method_limit)

@router.message(AdminStates.waiting_method_limit, F.text != 'বাতিল')
async def process_save_method_limit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    method_key = data.get('method_key')
    await state.clear()
    markup = await user_money_menu() if callable(user_money_menu) else user_money_menu
    try:
        raw_input = message.text.replace(')', '').strip()
        new_limit = float(raw_input)
        await asyncio.to_thread(update_setting, method_key, new_limit)
        await message.answer(f"<b>উত্তোলন লিমিট সফলভাবে পরিবর্তন করা হয়েছে: ৳{new_limit:.2f}</b>", reply_markup=markup)
    except ValueError:
        await message.answer("<b>ভুল ইনপুট! অনুগ্রহ করে সঠিক সংখ্যা দিয়ে পুনরায় চেষ্টা করুন।</b>", reply_markup=markup)

# -------------------------------------------------------------
# 🎁 ১৩. রেফারেল কমিশন ও ভিডিও গাইড সেট
# -------------------------------------------------------------
@router.message(F.text == 'রেফারেল কমিশন')
async def set_ref_commission(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer("নতুন রেফারেল কমিশন (%) লিখুন:", reply_markup=markup)
    await state.set_state(AdminStates.waiting_ref_comm)

@router.message(AdminStates.waiting_ref_comm, F.text != 'বাতিল')
async def process_save_ref_commission(message: types.Message, state: FSMContext):
    await state.clear()
    markup = await user_money_menu() if callable(user_money_menu) else user_money_menu
    try:
        comm = int(message.text.strip())
        await asyncio.to_thread(update_setting, 'ref_commission', comm)
        await message.answer("কমিশন আপডেট সফল!", reply_markup=markup)
    except Exception:
        await message.answer("ভুল ইনপুট!", reply_markup=markup)

@router.message(F.text == 'কাজের ভিডিও সেট')
async def set_video_guide(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    markup = await cancel_keyboard() if callable(cancel_keyboard) else cancel_keyboard
    await message.answer("নতুন ভিডিও লিংক পেস্ট করুন:", reply_markup=markup)
    await state.set_state(AdminStates.waiting_video_link)

@router.message(AdminStates.waiting_video_link, F.text != 'বাতিল')
async def process_save_video_link(message: types.Message, state: FSMContext):
    await state.clear()
    markup = await task_settings_menu() if callable(task_settings_menu) else task_settings_menu
    await asyncio.to_thread(update_setting, 'video_link', message.text.strip())
    await message.answer("ভিডিও লিংক আপডেট সফল!", reply_markup=markup)

# -------------------------------------------------------------
# ❌ ১৪. গ্লোবাল ক্যানসেল
# -------------------------------------------------------------
@router.message(F.text == 'বাতিল')
async def global_cancel_admin(message: types.Message, state: FSMContext):
    await state.clear()
    markup = await admin_menu() if callable(admin_menu) else admin_menu
    await message.answer("বাতিল করা হয়েছে।", reply_markup=markup)
